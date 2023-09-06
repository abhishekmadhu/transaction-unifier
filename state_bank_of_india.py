from abc import ABC, abstractmethod
from json import load


transaction_table_headers: list[str] = ['Txn Date', 'Value Date', 'Description', 'Ref No./Cheque No.','Debit', 'Credit', 'Balance']


class BankStatementParser(ABC):
    __unique_key: str

    def set_unique_key(self, key: str) -> None:
        self.__unique_key = key

    @property
    def unique_key(self) -> str:
        return self.__unique_key
    
    @abstractmethod
    def parse_summary() -> None:
        pass

    @abstractmethod
    def parse_transactions() -> None:
        pass
    

OLD_FILE_FORMAT = 'xls'
NEW_FILE_FORMAT = 'xlsx'
MAX_SUMMARY_ROWS = 20
MAX_TRANSACTION_ROWS = 1000

# It is a must to open the file in excel and "Save As" to the new format before parsing, as SBI sends corrupt files.
class SBI(BankStatementParser):
    summary_keys: list[str] = ['Account Number', 'Account Name', 'Branch', 'Balance', 'CIF', 'IFS']
    
    def __init__(self) -> None:
        super().__init__()
        
        self.summary = {}
        self.transactions = []

    def parse_summary(self, filepath:str='') -> None:
        
        if not filepath:
            print('Filepath is empty')
            return


        extension: str = filepath.split(sep='.')[-1]
        if extension == OLD_FILE_FORMAT:
            summary_information:dict = self.parse_summary_old_format(filepath=filepath)
        elif extension == NEW_FILE_FORMAT:
            summary_information:dict = self.parse_summary_new_format(filepath=filepath)
        else:
            return

        pass   
        
    def parse_summary_old_format(self, filepath:str) -> dict:
        if not filepath:
            return {}

        import xlrd

        workbook = None
        try:
            workbook = xlrd.open_workbook(filepath)
        except xlrd.XLRDError as e:
            print(e)
            print('You must open the file in excel and "Save As" to the new format before parsing, as SBI sends corrupt files.')
            return {}

        # the first and only sheet contains the transactions
        worksheet = workbook.sheet_by_index(0)

        rows = worksheet.nrows
        columns = worksheet.ncols

        summary_information: dict = {}
        for row in range(rows):
            cell_value = worksheet.cell_value(row, 0)
            for key in SBI.summary_keys:
                if key in cell_value:
                    summary_information[key] = worksheet.cell_value(row, 1)
                    break

        self.summary = summary_information
        return summary_information

    def parse_summary_new_format(self, filepath:str) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(filepath)
        sheet = wb.active

        # for utility, remove later
        self.wb = wb
        self.sheet = sheet
        
        summary_information: dict = {}

        for row in range(1, MAX_SUMMARY_ROWS):
            cell_value = sheet.cell(row=row, column=1).value
            for key in SBI.summary_keys:
                if key in cell_value:
                    summary_information[key] = sheet.cell(row=row, column=2).value
                    break

        self.summary = summary_information
        return summary_information

    def find_transaction_header_row(self, sheet) -> int:
        for row in range(1, MAX_TRANSACTION_ROWS):
            cell_value = sheet.cell(row=row, column=1).value
            if cell_value in transaction_table_headers:
                return row
        return 0

    def parse_transactions(self) -> list:
        transaction_start_row: int = self.find_transaction_header_row(sheet=self.sheet)
        transaction_end_row: int = self.sheet.max_row > MAX_TRANSACTION_ROWS and MAX_TRANSACTION_ROWS or self.sheet.max_row

        transactions = []

        if not self.sheet:
            return []
    
        # iterate over the rows and collect the data in a list of dictionaries
        for row in range(transaction_start_row+1, transaction_end_row):
            transaction: dict = {}
            for column in range(1, self.sheet.max_column):
                cell_value = self.sheet.cell(row=row, column=column).value
                if cell_value:
                    transaction[transaction_table_headers[column-1]] = cell_value

            transaction['unique_parent_key'] = self.unique_key

            # use the Txn Date and Description to create a unique_key
            transaction['unique_key'] = '-'.join([
                'SBI', 
                str(transaction.get('Txn Date', '').date() if transaction.get('Txn Date') else ''), 
                transaction['Description']
            ])

            # parse a category from the Description field
            # INB, UPI, IMPS, NEFT, ATM, POS, ECS, CHQ, CASH, FEE, TAX, INT, OTH
            

            transactions.append(transaction)

        self.transactions = transactions
        return transactions

    def build_unique_key(self) -> str:
        if not self.summary:
            return ''

        unique_key: str = '-'.join(['SBI', self.summary['IFS'], self.summary['Account Number']])
        return unique_key


# test
if __name__ == '__main__':
    sbi = SBI()
    results = sbi.parse_summary(filepath='samples/state_bank_of_india/sbi-statement.xlsx')
    unique_key = sbi.build_unique_key()
    transction_start = sbi.find_transaction_header_row(sheet=sbi.sheet)
    transactions = sbi.parse_transactions()
    pass